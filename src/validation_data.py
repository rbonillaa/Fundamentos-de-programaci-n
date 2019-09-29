from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from estudiante import Estudiante
import pandas as pd
import sys


def fillCell(cell,color,fill_type=None):
	fill = PatternFill(fgColor=color,fill_type=fill_type)
	cell.fill = fill

def convertirDatosNumericos(estudiante,row,columnas):
	for i,columna in enumerate(columnas):
		if columna == None:
			continue
		if "exam" in columna:
			for examen, temas in estudiante.examen.items():
				for tema, calif in temas.items():
					if calif and examen+tema == columna: 
						row[i].value = calif
		else:
			for actividad,dataD in estudiante.actividades.items():
				for item,calif in dataD.items():
					if calif and item==columna:
						row[i].value = calif

def validateDataSheet(filename,columnas,colores,errorsMessages,validaciones):
	header_rows = 2
	VACIO = errorsMessages[0]
	NO_REDONDEADO = errorsMessages[-1]
	TIPO_NO_NUMERICO = errorsMessages[-2]
	try:
		wb = load_workbook(filename)
		data_sheet = wb["DATA"]
		total_students = data_sheet.max_row - header_rows
		total_columns = data_sheet.max_column 

		print("Total estudiantes a revisar: "+str(total_students))
		total_errors = 0
		for row in data_sheet.iter_rows(min_row=header_rows+1, max_col=total_columns, max_row=data_sheet.max_row):
			data = []
			for cell in row:
				data.append(cell.value)
				fillCell(cell,colores[VACIO]) #Restore cell color
			estudiante = Estudiante(data, columnas)
			estudiante.convertir()
			convertirDatosNumericos(estudiante,row,columnas)
			#Luego de convertir datos numéricos, se valida
			errores = estudiante.validar(errorsMessages,validaciones)
			countErrorStudent = 0
			for error, campos in errores.items():
				color = colores[error]
				countErrorStudent += len(campos)
				for columna in campos:
					if columna in columnas:
						i = columnas.index(columna)
						if error==NO_REDONDEADO: #El error detectado es corregido por el script
							row[i].value = estudiante.calif_final[columna]
						else:
							fillCell(row[i],color,'solid') #Errores no corregidos se pintan
			
			if countErrorStudent:
				print('{:<9}{}'.format('FALLÓ',estudiante.nombre))
				total_errors += 1
			else:
				print('{:<9}{}'.format('OK',estudiante.nombre))
		if total_errors:
			print("RESULTADOS: Se detectaron errores en las calificaciones de ", total_errors," estudiantes.")
			print()
			print("Revise su archivo de calificaciones, y corrija los errores de las CELDAS RESALTADAS con los siguientes colores:")
			print("1. Naranja \t --> \t Dato requerido.")
			print("2. Azul \t --> \t Valor no válido, revise la cabecera del archivo modelo para las posibles opciones.")
			print("3. Amarillo \t --> \t Valor fuera del rango posible.")
			#print("4. Verde \t --> \t El dato no es numérico.")
			print("="*150)
			print("**OJO: Los errores tipo dato_no_redondeado y dato_no_numerico fueron corregidos por el script")
		else:
			print("="*150)
			print('¡Felicidades! Tu archivo de excel está listo para ser enviado')
			print('Por favor enviar a rabonilla@fiec.espol.edu.ec con el subject Estadísticas P##-2019I')
            print('No te olvides de renombrar tu archivo excel como estadisticas-P##-2019I.xlsx')
		print("="*150)
		print("Copyright (c) 2018 eslozano")
		print('All Rights Reserved :)')
		wb.save(filename)
	except IOError as e:
		print("ERROR: No." , e.errno , e )
		if e.errno == 2:
			print("El archivo ",filename," no se encuentra en la carpeta del script.")
		if e.errno == 13:
			print("El archivo ",filename," se encuentra abierto.")


def getColumns(filename):
	wb = load_workbook(filename)
	data_sheet = wb["DATA"]
	total_columns = data_sheet.max_column

	rows = []

	for row in data_sheet.iter_rows(min_row=0, max_col=total_columns, max_row=2):
		column_data = []
		for cell in row:
			column_data.append(cell.value)
		rows.append(column_data)

	wb.close()
	return tuple(rows)


#---- Main ------#

if (len(sys.argv)<2):
	print("Error, ingrese el nombre del archivo")
	print("$python3 validation_data.py archivo.xlsx")
	exit()

filename = sys.argv[1]
print("**",filename,"**")


OPCION_NO_VALIDA = 'dato_no_valido'
VACIO = 'dato_vacio'
FUERA_DE_RANGO = 'dato_fuera_de_rango'
NO_REDONDEADO = 'dato_no_redondeado'
TIPO_NO_NUMERICO = 'dato_no_numerico'

#Los colores en los que se resaltarán las celdas según los errores
colors = {
	VACIO:"ffd8bf", #naranja 
	OPCION_NO_VALIDA: '6aa2fc', #azul 
	FUERA_DE_RANGO: 'fffa00', #amarillo
	TIPO_NO_NUMERICO: '96D701', #verde
	NO_REDONDEADO: 'ff7ff4' #rosado
}

errorsMessages= [VACIO,OPCION_NO_VALIDA,FUERA_DE_RANGO,TIPO_NO_NUMERICO,NO_REDONDEADO]

#df = pd.read_excel(filename)
#columnas = list(df.columns)
#validations = list(df.iloc[0])

columnas,validations=getColumns(filename)

"""
se crea un diccionario con el siguiente formato
"""

validaciones = {
	"genero": ["F","M"],
	"veces_tomadas":[2,3],
	"calif_final": 100,
	"sustent": 1,
	"proyecto":{"1er_proyecto":20,"2do_proyecto":20,"3er_proyecto":25},
	"lecciones":10
}

# se agregan las preguntas del examen
examen_dict = {}
exam_list = ["1er_exam_","2do_exam_","3er_exam_"]
for exam in exam_list:
	q = {}
	for i,c in enumerate(columnas):
		print(c)
		#if c == None:
			#continue;
		if c != None and exam in c:
			q[c] = validations[i]
	examen_dict[exam]=q


validaciones["examen"]=examen_dict

validateDataSheet(filename,columnas,colors,errorsMessages, validaciones)
